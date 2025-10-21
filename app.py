from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import datetime
import os

app = Flask(__name__)

# Đọc data
DATA_FILE = os.path.join('static', 'data.csv')

@app.route('/')
def index():
    # Đọc CSV KHÔNG CÓ HEADER
    df = pd.read_csv(DATA_FILE, header=None, names=['Tên', 'Mã hội viên', 'Quyền'], encoding='utf-8')
    members = df.to_dict(orient='records')
    return render_template('index.html', members=members)

@app.route('/export', methods=['POST'])
def export():
    try:
        selected = request.json.get('selected', [])
        exam_code = request.json.get('exam_code', '').strip()
        
        if not selected:
            return jsonify({'error': 'Chưa chọn học viên nào'}), 400
        if not exam_code:
            return jsonify({'error': 'Chưa nhập mã kỳ thi'}), 400
        
        # Đọc CSV KHÔNG CÓ HEADER
        df = pd.read_csv(DATA_FILE, header=None, names=['Tên', 'Mã hội viên', 'Quyền'], encoding='utf-8')
        
        # Lọc theo mã được chọn và GIỮ NGUYÊN THỨ TỰ CHỌN
        df_sel_list = []
        for code in selected:
            row = df[df['Mã hội viên'] == code]
            if not row.empty:
                df_sel_list.append(row.iloc[0])
        
        if not df_sel_list:
            return jsonify({'error': 'Không tìm thấy học viên'}), 400
        
        # Tính cấp đăng ký dự thi (chỉ trả về SỐ, bỏ chữ "Cấp")
        def calculate_cap_dang_ky(quyen):
            quyen_str = str(quyen).strip()
            
            # Xử lý "Cấp X" -> trả về số X-1
            if quyen_str.startswith('Cấp'):
                try:
                    cap_hien_tai = int(quyen_str.replace('Cấp', '').strip())
                    cap_dang_ky = cap_hien_tai - 1
                    return cap_dang_ky  # Chỉ trả về số, không có chữ "Cấp"
                except:
                    return quyen_str
            
            # Các trường hợp khác (Đẳng, GV) giữ nguyên
            else:
                return quyen_str
        
        # Tạo workbook mới
        wb = Workbook()
        ws = wb.active
        ws.title = "DST"
        
        # Thiết lập độ rộng cột
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 28
        
        # Tiêu đề chính (dòng 1-2) - MERGE CELLS
        ws.merge_cells('A1:F2')
        title_cell = ws['A1']
        title_cell.value = 'DANH SÁCH ĐĂNG KÝ THAM DỰ THI THĂNG CẤP ĐAI TAEKWONDO CLB_01102'
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Header (dòng 3)
        headers = ['STT', 'Mã kỳ thi', 'Mã Đơn vị', 'Mã CLB', 'Mã hội viên', 'Cấp đẳng đăng ký dự thi']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Ghi dữ liệu (từ dòng 4) - THEO THỨ TỰ CHỌN
        for idx, row_data in enumerate(df_sel_list, start=1):
            current_row = 3 + idx
            
            # STT
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=1).font = Font(name='Arial', size=11)
            
            # Mã kỳ thi
            ws.cell(row=current_row, column=2, value=exam_code)
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=2).font = Font(name='Arial', size=11)
            
            # Mã đơn vị
            ws.cell(row=current_row, column=3, value='TNIN')
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=3).font = Font(name='Arial', size=11)
            
            # Mã CLB
            ws.cell(row=current_row, column=4, value='CLB_01102')
            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=4).font = Font(name='Arial', size=11)
            
            # Mã hội viên
            ws.cell(row=current_row, column=5, value=row_data['Mã hội viên'])
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=5).font = Font(name='Arial', size=11)
            
            # Cấp đăng ký dự thi
            cap_dang_ky = calculate_cap_dang_ky(row_data['Quyền'])
            ws.cell(row=current_row, column=6, value=cap_dang_ky)
            ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=6).font = Font(name='Arial', size=11)
            
            # Border cho tất cả các ô
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Lưu vào BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Tên file với timestamp
        filename = f"DST_{exam_code}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output, 
            as_attachment=True, 
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        app.logger.error(f"Lỗi tạo Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi tạo file Excel: {str(e)}'}), 500

if __name__ == '__main__':
    # Lấy PORT từ môi trường (cho Render)
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)